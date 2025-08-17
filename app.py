import io
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd
import streamlit as st
import matplotlib.pyplot as plt

# ---------------------------
# Page setup
# ---------------------------
st.set_page_config(page_title="Data Quality Reporter", layout="wide")

st.title("📊 Data Quality Reporter")
st.caption("Upload a CSV/Excel file → get a TXT + Excel report → scan issues with a heatmap.")

# ---------------------------
# Core helpers (profiling-lite)
# ---------------------------
def _missing_table(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame(columns=["Missing", "Missing %"])
    missing = df.isna().sum()
    pct = (missing / len(df)) * 100 if len(df) else 0
    out = (
        pd.DataFrame({"Missing": missing, "Missing %": pct})
        .sort_values("Missing", ascending=False)
    )
    return out[out["Missing"] > 0]

def _dtype_table(df: pd.DataFrame) -> pd.DataFrame:
    vc = df.dtypes.value_counts(dropna=False).rename_axis("dtype").reset_index(name="count")
    return vc

def _dtype_columns_map(df: pd.DataFrame) -> pd.DataFrame:
    rows = []
    for dt, cols in df.dtypes.groupby(df.dtypes).groups.items():
        rows.append({"dtype": str(dt), "columns": ", ".join(list(cols))})
    return pd.DataFrame(rows).sort_values("dtype").reset_index(drop=True)

def _numeric_summary(df: pd.DataFrame) -> pd.DataFrame:
    num = df.select_dtypes(include=[np.number])
    if num.empty:
        return pd.DataFrame()
    desc = num.describe(percentiles=[0.25, 0.5, 0.75]).T
    desc = desc.rename(columns={"25%": "Q1", "50%": "Median", "75%": "Q3"})
    desc["skew"] = num.skew(numeric_only=True)
    desc["kurt"] = num.kurt(numeric_only=True)
    return desc.reset_index().rename(columns={"index": "column"})

def _categorical_summary(df: pd.DataFrame) -> pd.DataFrame:
    cat_cols = df.select_dtypes(exclude=[np.number]).columns
    rows = []
    for c in cat_cols:
        s = df[c]
        top, freq = (None, 0)
        if len(s.dropna()) > 0:
            vc = s.value_counts(dropna=True)
            top = vc.index[0]
            freq = int(vc.iloc[0])
        rows.append({
            "column": c,
            "non_null_count": s.notna().sum(),
            "unique": s.nunique(dropna=True),
            "top_value": top,
            "top_value_freq": freq,
        })
    return pd.DataFrame(rows).sort_values("column").reset_index(drop=True)

def _iqr_outliers(df: pd.DataFrame, max_index_list: int = 25) -> pd.DataFrame:
    num = df.select_dtypes(include=[np.number])
    rows = []
    n = len(df)
    if num.empty or n == 0:
        return pd.DataFrame(columns=["column", "outliers_count", "outliers_pct", "sample_indices"])
    for c in num.columns:
        x = num[c].dropna()
        if x.empty:
            continue
        q1 = x.quantile(0.25)
        q3 = x.quantile(0.75)
        iqr = q3 - q1
        lower = q1 - 1.5 * iqr
        upper = q3 + 1.5 * iqr
        mask = (df[c] < lower) | (df[c] > upper)
        idx = df.index[mask].tolist()
        rows.append({
            "column": c,
            "outliers_count": len(idx),
            "outliers_pct": (len(idx) / n) * 100,
            "sample_indices": idx[:max_index_list],
        })
    out = pd.DataFrame(rows).sort_values("outliers_count", ascending=False)
    return out[out["outliers_count"] > 0].reset_index(drop=True)

def _duplicate_preview(df: pd.DataFrame, max_rows: int = 200) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()
    dups = df[df.duplicated(keep=False)]
    if dups.empty:
        return pd.DataFrame()
    return dups.head(max_rows)

# ---------------------------
# Excel writer (openpyxl)
# ---------------------------
def _write_conditional_formats(ws, max_row, max_col, missing_pct_col_letter: Optional[str]):
    from openpyxl.formatting.rule import ColorScaleRule, CellIsRule
    from openpyxl.styles import Font

    rng = f"A2:{chr(64 + max_col)}{max_row}"
    ws.conditional_formatting.add(
        rng,
        ColorScaleRule(start_type='min', mid_type='percentile', mid_value=50, end_type='max')
    )

    if missing_pct_col_letter:
        ws.conditional_formatting.add(
            f"{missing_pct_col_letter}2:{missing_pct_col_letter}{max_row}",
            CellIsRule(operator='greaterThanOrEqual', formula=['10'], stopIfTrue=True, font=Font(bold=True))
        )

def _add_missing_chart(wb, ws_name: str):
    from openpyxl.chart import BarChart, Reference
    ws = wb[ws_name]
    max_row = ws.max_row
    if max_row < 3:
        return
    cats = Reference(ws, min_col=1, min_row=2, max_row=max_row)  # Column names
    vals = Reference(ws, min_col=3, min_row=2, max_row=max_row)  # Missing %
    chart = BarChart()
    chart.title = "Missing Percentage by Column"
    chart.y_axis.title = "Missing %"
    chart.x_axis.title = "Column"
    chart.add_data(vals, titles_from_data=False)
    chart.set_categories(cats)
    ws.add_chart(chart, "E2")

def data_quality_report_to_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # Overview
        overview_rows = [
            ["Rows", len(df)],
            ["Columns", df.shape[1]],
            ["Duplicate rows", int(df.duplicated().sum())],
        ]
        dtype_counts = _dtype_table(df)
        overview_df = pd.DataFrame(overview_rows, columns=["Metric", "Value"])
        overview_df.to_excel(writer, sheet_name="Overview", index=False)
        dtype_counts.to_excel(writer, sheet_name="Overview", index=False, startrow=len(overview_rows)+2)

        # Missing
        missing_df = _missing_table(df)
        if missing_df.empty:
            missing_df = pd.DataFrame(columns=["Missing", "Missing %"])
        missing_df.index.name = "column"
        missing_df = missing_df.reset_index()
        missing_df.to_excel(writer, sheet_name="Missing Values", index=False)

        # Data Types map
        dtype_cols_df = _dtype_columns_map(df)
        dtype_cols_df.to_excel(writer, sheet_name="Data Types", index=False)

        # Numeric / Categorical
        num_sum = _numeric_summary(df)
        if num_sum.empty:
            num_sum = pd.DataFrame(columns=["column","count","mean","std","min","Q1","Median","Q3","max","skew","kurt"])
        num_sum.to_excel(writer, sheet_name="Numeric Summary", index=False)

        cat_sum = _categorical_summary(df)
        cat_sum.to_excel(writer, sheet_name="Categorical Summary", index=False)

        # Outliers, Duplicates, Sample
        outliers = _iqr_outliers(df)
        if outliers.empty:
            outliers = pd.DataFrame(columns=["column","outliers_count","outliers_pct","sample_indices"])
        outliers.to_excel(writer, sheet_name="Outliers (IQR)", index=False)

        dups = _duplicate_preview(df)
        dups.to_excel(writer, sheet_name="Duplicates", index=False)

        df.head(100).to_excel(writer, sheet_name="Sample Data", index=False)

        # Styling / chart
        wb = writer.book
        ws_miss = wb["Missing Values"]
        _write_conditional_formats(ws_miss, ws_miss.max_row, ws_miss.max_column, missing_pct_col_letter="C")
        _add_missing_chart(wb, "Missing Values")

        for sheet in ["Numeric Summary", "Categorical Summary", "Outliers (IQR)"]:
            ws = wb[sheet]
            _write_conditional_formats(ws, ws.max_row, ws.max_column, None)

    output.seek(0)
    return output.read()

def data_quality_report_txt(df: pd.DataFrame) -> str:
    report = []
    report.append("DATA QUALITY REPORT")
    report.append("=" * 70)
    report.append(f"Dataset shape: {df.shape[0]} rows × {df.shape[1]} columns\n")

    miss = _missing_table(df)
    report.append("MISSING VALUES:")
    if miss.empty:
        report.append("  ✓ No missing values found\n")
    else:
        for _, r in miss.iterrows():
            report.append(f"  {r['column']}: {int(r['Missing'])} ({r['Missing %']:.1f}%)")
        report.append("")

    dups = int(df.duplicated().sum())
    report.append(f"DUPLICATE ROWS: {dups}\n")

    report.append("DATA TYPES SUMMARY:")
    for _, row in _dtype_table(df).iterrows():
        report.append(f"  {row['dtype']}: {row['count']} columns")
    report.append("")

    out = _iqr_outliers(df)
    report.append("NUMERIC OUTLIERS (IQR method):")
    if out.empty:
        report.append("  ✓ No potential outliers detected by IQR\n")
    else:
        for _, r in out.iterrows():
            report.append(f"  {r['column']}: {int(r['outliers_count'])} ({r['outliers_pct']:.1f}%)")
    report.append("")

    return "\n".join(report)

# ---------------------------
# Issues Heatmap (per-column)
# ---------------------------
def compute_issue_matrix(df: pd.DataFrame) -> pd.DataFrame:
    """
    Build an issue matrix with rows as metrics and columns as df columns:
      - Missing %
      - Outlier % (numeric only)
      - Zero % (numeric only)
      - Negative % (numeric only)
      - Constant % (share of most frequent value)
      - Cardinality Ratio (unique/non-null)
    """
    n = len(df)
    cols = df.columns
    mat = pd.DataFrame(index=["Missing %","Outlier %","Zero %","Negative %","Constant %","Cardinality Ratio"],
                       columns=cols, dtype=float)

    # Missing %
    if n > 0:
        mat.loc["Missing %"] = (df.isna().sum() / n) * 100
    else:
        mat.loc["Missing %"] = 0.0

    # Outliers, Zero, Negative for numeric
    num_cols = df.select_dtypes(include=[np.number]).columns
    mat.loc["Outlier %"] = 0.0
    mat.loc["Zero %"] = 0.0
    mat.loc["Negative %"] = 0.0

    if n > 0 and len(num_cols) > 0:
        # Outlier %
        outliers = _iqr_outliers(df)
        if not outliers.empty:
            for _, r in outliers.iterrows():
                c = r["column"]
                if c in mat.columns:
                    mat.loc["Outlier %", c] = float(r["outliers_pct"])

        # Zero % & Negative %
        for c in num_cols:
            s = df[c]
            mat.loc["Zero %", c] = (s.eq(0).sum() / n) * 100
            mat.loc["Negative %", c] = (s.lt(0).sum() / n) * 100

    # Constant % & Cardinality ratio for all
    for c in cols:
        s = df[c]
        non_null = s.dropna()
        if len(non_null) == 0:
            mat.loc["Constant %", c] = 0.0
            mat.loc["Cardinality Ratio", c] = 0.0
        else:
            vc = non_null.value_counts()
            mat.loc["Constant %", c] = (vc.iloc[0] / len(non_null)) * 100
            mat.loc["Cardinality Ratio", c] = non_null.nunique() / len(non_null)

    return mat.fillna(0.0)

def plot_issue_heatmap(mat: pd.DataFrame):
    """Render a simple matplotlib heatmap (no custom colors as per guideline)."""
    fig, ax = plt.subplots(figsize=(max(8, 0.3*mat.shape[1]), 4.5))
    im = ax.imshow(mat.values, aspect="auto")
    ax.set_xticks(np.arange(mat.shape[1]))
    ax.set_xticklabels(mat.columns, rotation=45, ha="right")
    ax.set_yticks(np.arange(mat.shape[0]))
    ax.set_yticklabels(mat.index)
    ax.set_title("Issues Heatmap (higher = more risk)")
    plt.tight_layout()
    return fig

# ---------------------------
# File reading utilities
# ---------------------------
@st.cache_data(show_spinner=False)
def load_csv(file, encoding_try: str = "utf-8", sep: str = ",") -> pd.DataFrame:
    try:
        return pd.read_csv(file, sep=sep, encoding=encoding_try)
    except UnicodeDecodeError:
        file.seek(0)
        return pd.read_csv(file, sep=sep, encoding="latin-1")

@st.cache_data(show_spinner=False)
def load_excel(file, sheet: Optional[str] = None) -> pd.DataFrame:
    try:
        return pd.read_excel(file, sheet_name=sheet)
    except ValueError:
        # If the provided sheet doesn't exist, fallback to first sheet
        xls = pd.ExcelFile(file)
        return pd.read_excel(file, sheet_name=xls.sheet_names[0])

def infer_and_cast_numerics(df: pd.DataFrame) -> pd.DataFrame:
    df2 = df.copy()
    for c in df2.columns:
        if df2[c].dtype == object:
            # Try to coerce numeric-like strings
            coerced = pd.to_numeric(df2[c].astype(str).str.replace(",","", regex=False), errors="coerce")
            # Only adopt if we get a meaningful number of non-nulls
            if coerced.notna().sum() >= max(5, 0.2*len(df2)):
                df2[c] = coerced
    return df2

# ---------------------------
# Sidebar: controls
# ---------------------------
st.sidebar.header("Upload & Options")
file = st.sidebar.file_uploader("Upload CSV or Excel", type=["csv","xlsx","xls"], accept_multiple_files=False)
cast_numeric = st.sidebar.checkbox("Auto-convert numeric-looking text", value=True)

csv_sep = st.sidebar.text_input("CSV delimiter (if CSV)", value=",")
sheet_name = st.sidebar.text_input("Excel sheet name (optional)", value="")

show_preview = st.sidebar.checkbox("Show data preview (head)", value=True)
preview_rows = st.sidebar.number_input("Preview rows", min_value=5, max_value=200, value=20, step=5)

# ---------------------------
# Main logic
# ---------------------------
df = None
if file is not None:
    name = file.name.lower()
    if name.endswith(".csv"):
        df = load_csv(file, sep=csv_sep or ",")
    else:
        df = load_excel(file, sheet=sheet_name if sheet_name.strip() else None)

    if cast_numeric:
        df = infer_and_cast_numerics(df)

    st.success(f"Loaded **{file.name}** — {df.shape[0]} rows × {df.shape[1]} columns.")
    if show_preview:
        st.subheader("🔎 Data preview")
        st.dataframe(df.head(int(preview_rows)), use_container_width=True)

    # Actions
    colA, colB, colC = st.columns([1,1,1])
    with colA:
        gen = st.button("🧾 Generate Reports")
    with colB:
        show_heatmap = st.toggle("Show Issues Heatmap", value=True)
    with colC:
        show_dups = st.toggle("Show Duplicate Rows Preview", value=False)

    if gen:
        with st.spinner("Building TXT + Excel reports..."):
            txt = data_quality_report_txt(df)
            excel_bytes = data_quality_report_to_excel(df)

            st.session_state["dqr_txt"] = txt
            st.session_state["dqr_xlsx"] = excel_bytes

        st.success("Reports generated.")

    # Downloads (one-click after generate)
    if "dqr_txt" in st.session_state and "dqr_xlsx" in st.session_state:
        dl1, dl2 = st.columns([1,1])
        with dl1:
            st.download_button(
                "⬇️ Download TXT Report",
                data=st.session_state["dqr_txt"].encode("utf-8"),
                file_name="data_quality_report.txt",
                mime="text/plain",
            )
        with dl2:
            st.download_button(
                "⬇️ Download Excel Report",
                data=st.session_state["dqr_xlsx"],
                file_name="data_quality_report.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

    # Heatmap
    if show_heatmap:
        st.subheader("🌡️ Issues Heatmap")
        mat = compute_issue_matrix(df)
        fig = plot_issue_heatmap(mat)
        st.pyplot(fig, clear_figure=True)

        # Export heatmap
        buf = io.BytesIO()
        fig.savefig(buf, format="png", bbox_inches="tight")
        buf.seek(0)
        st.download_button(
            "⬇️ Download Heatmap (PNG)",
            data=buf,
            file_name="issues_heatmap.png",
            mime="image/png"
        )

    # Duplicates preview
    if show_dups:
        st.subheader("🧭 Duplicate Rows (preview)")
        dups = _duplicate_preview(df)
        if dups.empty:
            st.info("No duplicate rows detected.")
        else:
            st.dataframe(dups, use_container_width=True)

else:
    st.info("Upload a CSV or Excel file in the sidebar to get started.")
    st.markdown(
        "- CSV tip: adjust the delimiter if your file uses `;` or `\\t`.\n"
        "- Excel tip: specify a sheet name if needed."
    )

# ---------------------------
# Footer
# ---------------------------
st.caption("© Data Quality Reporter — lightweight profiling for fast, repeatable checks.")
